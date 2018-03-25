# coding: utf-8


#############################
# import
import sys
import codecs
import io
import re
import openpyxl



#############################
# グローバル

# 局数
station_num=0

# 局名
station_name=[]

# 局規模
station_size=[]

# 管理番号
station_no=[]

# 放送局有無
braodcast_exit00 = []
braodcast_exit01 = []
braodcast_exit02 = []
braodcast_exit03 = []
braodcast_exit04 = []
braodcast_exit05 = []
braodcast_exit06 = []
braodcast_exit07 = []
braodcast_exit08 = []
braodcast_exit09 = []
braodcast_exit10 = []
braodcast_exit11 = []
braodcas_exit = [braodcast_exit00, braodcast_exit01, braodcast_exit02, braodcast_exit03, braodcast_exit04, braodcast_exit05, 
braodcast_exit06,  braodcast_exit07, braodcast_exit08, braodcast_exit09, braodcast_exit10, braodcast_exit11]

# 放送局数
braodcast_idx = 0

# 開局日
station_opendate=[]

#############################
# 初期化
def initGlobal():
	# 局数
	global station_num
	station_num=0
	# 放送局数
	global braodcast_idx
	braodcast_idx=0

	# 局名
	station_name.clear()

	# 局規模
	station_size.clear()

	# 管理番号
	station_no.clear()

	# 放送局有無
	braodcast_exit00.clear()
	braodcast_exit01.clear()
	braodcast_exit02.clear()
	braodcast_exit03.clear()
	braodcast_exit04.clear()
	braodcast_exit05.clear()
	braodcast_exit06.clear()
	braodcast_exit07.clear()
	braodcast_exit08.clear()
	braodcast_exit09.clear()
	braodcast_exit10.clear()
	braodcast_exit11.clear()

	# 開局日
	station_opendate.clear()

#############################
# 局名収集
def analyze_station(file_p):
	for line in file_p:
		if '局所' in line:
			print(len(station_name))
			print(station_name)
			return len(station_name);
		elif line.strip() != '':
			station_name.append(line.strip())
			#print(line.strip());

#############################
# 放送局名収集
def analyze_broadcast(file_p):
	braodcast=[]
	for line in file_p:
		if '◎' in line:
			print(len(braodcast))
			print(braodcast)
			return;
		elif line.strip() != '':
			braodcast.append(line.strip())
			#print(line.strip())

#############################
# 局規模収集
def analyze_stationsize(file_p, station_num):
	items = 1
	while items < station_num:
		line = file_p.readline()
		if line.strip() != '':
			station_size.append(line.strip())
			items = items + 1
	print(len(station_size))
	print(station_size)

#############################
# 局管理番号収集
def analyze_station_no(file_p, station_num):
	items = 1
	while items < station_num:
		line = file_p.readline()
		if line.strip() != '':
			station_no.append(line.strip())
			items = items + 1
	print(len(station_no))
	print(station_no)

#############################
# 局規有無
def analyze_braodcast_exist(file_p, braodcast_idx, station_num):
	items = 1
	while items < station_num:
		line = file_p.readline()
		if line.strip() != '':
			braodcas_exit[braodcast_idx].append(line.strip())
			items = items + 1
	print(len(braodcas_exit[braodcast_idx]))
	print(braodcas_exit[braodcast_idx])

#############################
# 開局日
def analyze_station_opnedate(file_p, station_num):
	opendate_regex = re.compile(r'\d{4}年\d{2}月\d{2}日')
	items = 1
	while items < station_num:
		line = file_p.readline()
		opendate = opendate_regex.match(line)
		if opendate:
			station_opendate.append(opendate.group())
			items = items + 1
	print(len(station_opendate))
	print(station_opendate)


#############################
# テキスト変換情報解析
def analyze_text(converted_pdf_file):
	global station_num
	global braodcast_idx

	opendate_regex = re.compile(r'\d{4}年\d{2}月\d{2}日')
	
	fp_in = codecs.open(converted_pdf_file, 'r', 'utf-8');
	for line in fp_in:
		#都道府県の認識
		if '都道府県' in line:
			pref = line.split()
			print(pref[1])
		
		# 局名の認識
		if re.match('局名$', line):
			station_num = analyze_station(fp_in)
			
		# 放送局名の認識
		#if re.match('管理番号$', line):
		#	analyze_broadcast(fp_in)

		# 局規模
		if re.match('◎|◉|･', line):
			station_size.append(line.strip())
			analyze_stationsize(fp_in, station_num)
		
		# 管理番号
		if re.match(r'\d{6}', line):
			station_no.append(line.strip())
			analyze_station_no(fp_in, station_num)
		
		# 放送局有無
		if re.match('○|-', line):
			braodcas_exit[braodcast_idx].append(line.strip())
			analyze_braodcast_exist(fp_in, braodcast_idx, station_num)
			braodcast_idx = braodcast_idx + 1
		
		# 開設時期
		opendate = opendate_regex.match(line)
		if opendate:
			station_opendate.append(opendate.group())
			analyze_station_opnedate(fp_in, station_num)

#############################
# EXCELに吐き出す
def write_station_Excel(output_path):
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title='Station List'
	
	print(station_num)
	for row_num in range(station_num):
		sheet.cell(row=row_num+1, column=2).value = station_name[row_num]
		sheet.cell(row=row_num+1, column=3).value = station_no[row_num]
		
		for broad_idx in range(braodcast_idx):
			sheet.cell(row=row_num+1, column=4+broad_idx).value = braodcas_exit[broad_idx][row_num]
	
	wb.save(output_path)




#############################
# EXCELに吐き出す
def txt2Excel(input_path, output_path):
	
	initGlobal()
	
	# TXTを解析する"
	analyze_text(input_path)

	# EXCELに吐き出す
	write_station_Excel(output_path)


